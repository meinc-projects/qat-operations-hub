"""Export the most recent audit report to a formatted Excel file."""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data_dir = Path(__file__).resolve().parent.parent / "data"
reports = sorted(data_dir.glob("audit_report_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
if not reports:
    print("No audit reports found.")
    sys.exit(1)

report_path = reports[0]
print(f"Using: {report_path.name}")

with open(report_path, "r", encoding="utf-8") as f:
    report = json.load(f)

run_id = report.get("run_id", "unknown")[:8]
run_date = report.get("run_date", "")[:10]
summary = report.get("summary", {})
issues = report.get("issues", {})
total = report.get("total_deals_scanned", 0)
succeeded = report.get("successfully_processed", 0)

wb = Workbook()
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
title_font = Font(bold=True, size=14)

# --- Summary sheet ---
ws = wb.active
ws.title = "Summary"
ws.append(["QAT Operations Hub — Audit Report"])
ws["A1"].font = title_font
ws.append([])
ws.append(["Run Date", run_date])
ws.append(["Run ID", report.get("run_id", "")])
ws.append(["Total Deals Scanned", total])
ws.append(["Successfully Processed", succeeded])
ws.append(["Failed / Skipped", total - succeeded])
ws.append([])
ws.append(["COMPLIANCE FLAGS", "Count", "% of Total"])
ws.append(["Missing DMV Paperwork", summary.get("missing_dmv_paperwork", 0),
           f"{summary.get('missing_dmv_paperwork', 0) / max(total, 1) * 100:.0f}%"])
ws.append(["Missing Reg/Registration File", summary.get("missing_reg_file", 0),
           f"{summary.get('missing_reg_file', 0) / max(total, 1) * 100:.0f}%"])
ws.append(["No Attachments At All", summary.get("no_attachment", 0),
           f"{summary.get('no_attachment', 0) / max(total, 1) * 100:.0f}%"])
ws.append([])
ws.append(["DATA QUALITY", "Count", "% of Total"])
ws.append(["OCR Failed", summary.get("ocr_failed", 0),
           f"{summary.get('ocr_failed', 0) / max(total, 1) * 100:.0f}%"])
ws.append(["Invalid VIN", summary.get("invalid_vin", 0),
           f"{summary.get('invalid_vin', 0) / max(total, 1) * 100:.0f}%"])
ws.append(["Orphan Deal (no contact)", summary.get("orphan_deal", 0),
           f"{summary.get('orphan_deal', 0) / max(total, 1) * 100:.0f}%"])
ws.append(["Missing Email", summary.get("missing_email", 0),
           f"{summary.get('missing_email', 0) / max(total, 1) * 100:.0f}%"])
for row in ws["A9:C9"]:
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill
for row in ws["A14:C14"]:
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12


def add_issue_sheet(wb, name, issue_list, headers):
    ws = wb.create_sheet(title=name[:31])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for item in issue_list:
        ws.append([
            item.get("deal_id", ""),
            item.get("deal_name", ""),
            item.get("contact_name", ""),
            item.get("issue_details", ""),
            "",
            "",
        ])
    for i, w in enumerate([18, 50, 25, 60, 18, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


headers = ["Deal ID", "Deal Name", "Contact", "Files On Record", "Follow-Up Status", "Notes"]
add_issue_sheet(wb, "Missing DMV Paperwork", issues.get("missing_dmv_paperwork", []), headers)
add_issue_sheet(wb, "Missing Reg File", issues.get("missing_reg_file", []), headers)
add_issue_sheet(wb, "No Attachments", issues.get("no_attachment", []), headers)
add_issue_sheet(wb, "OCR Failed", issues.get("ocr_failed", []),
                ["Deal ID", "Deal Name", "Contact", "Error Details", "Follow-Up Status", "Notes"])
add_issue_sheet(wb, "Orphan Deals", issues.get("orphan_deal", []),
                ["Deal ID", "Deal Name", "Contact", "Details", "Follow-Up Status", "Notes"])

out_path = data_dir / f"audit_report_{run_id}.xlsx"
wb.save(out_path)
print(f"\nExcel report saved: {out_path}")
print(f"Upload to Google Drive: https://drive.google.com/drive/folders/1dmXUCxl1fUtuTURwvTvYp5JEz2EKPP9a")
